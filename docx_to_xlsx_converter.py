import flet as ft
import docx
import re
from openpyxl import Workbook

def main(page: ft.Page):
    page.title = "docx to xlsx converter"
    page.window_height=650
    page.window_width=750
    # Variable to store the path of the selected file
    selected_docx_path = None

    def on_dialog_result(e: ft.FilePickerResultEvent):
        nonlocal selected_docx_path
        if e.files:
            # If a file was selected, save the file path to a variable
            selected_docx_path = e.files[0].path
            # Update the text on the page to show the selected path
            docxfile_path_text.value = f"Выбранный файл: {selected_docx_path}"
            page.update()
        else:
            docxfile_path_text.value = "Файл не выбран"
            page.update()

    # Create a FilePicker instance
    docxfile_picker = ft.FilePicker(on_result=on_dialog_result)

    # Add FilePicker to the page (necessary for operation)
    page.overlay.append(docxfile_picker)

    # Create a text field to display the selected path
    docxfile_path_text = ft.Text(value="Файл не выбран", color=ft.colors.RED)

    # Create a button to call the file selection dialog
    pick_docxfile_button = ft.ElevatedButton(
        "Выбрать файл",
        icon=ft.icons.UPLOAD_FILE,
        on_click=lambda _: docxfile_picker.pick_files(allow_multiple=False),
    )

    page.add(ft.Text('ВНИМАНИЕ! В ДОКУМЕНТ EXEL БУДУТ ДОБАВЛЕНЫ НОВЫЕ ДАННЫЕ, РЕКОМЕНДУЕТСЯ ИСПОЛЬЗОВАТЬ ПУСТОЙ ДОКУМЕНТ', color=ft.colors.RED))
    page.add(ft.Text('Выберите файл .docx из которого будет составлен список элементов'))
    # Add the button and text field to the page
    page.add(pick_docxfile_button, docxfile_path_text)
    
    selected_xlsx_path = None

    def on_dialog_result(e: ft.FilePickerResultEvent):
        nonlocal selected_xlsx_path
        if e.files:
            # If a file was selected, save the file path to a variable
            selected_xlsx_path = e.files[0].path
            # Update the text on the page to show the selected path
            xlsxfile_path_text.value = f"Выбранный файл: {selected_xlsx_path}"
            page.update()
        else:
            xlsxfile_path_text.value = "Файл не выбран"
            page.update()

    # Create a FilePicker instance
    xlsxfile_picker = ft.FilePicker(on_result=on_dialog_result)

    # Add FilePicker to the page (necessary for operation)
    page.overlay.append(xlsxfile_picker)

    # Create a text field to display the selected path
    xlsxfile_path_text = ft.Text(value="Файл не выбран", color=ft.colors.RED)

    # Create a button to call the file selection dialog
    pick_xlsxfile_button = ft.ElevatedButton(
        "Выбрать файл",
        icon=ft.icons.UPLOAD_FILE,
        on_click=lambda _: xlsxfile_picker.pick_files(allow_multiple=False),
    )

    page.add(ft.Text('Выберите файл .xlsx куда нажно вставить список по 1 элементу'))
    # Add the button and text field to the page
    page.add(pick_xlsxfile_button, xlsxfile_path_text)

    def extract_text_from_docx(file_path):
        doc = docx.Document(file_path)
        full_text = []
        for para in doc.paragraphs:
                if word_checkbox.value==0 and number_checkbox.value==0:
                    full_text.append(para.text)
                elif word_checkbox.value==0 and number_checkbox.value==1:
                    full_text = re.sub(r'\b[^\W\d_]+\b','', para.text)
                elif word_checkbox.value==1 and number_checkbox.value==0:
                    full_text = re.sub(r'\d+','', para.text)
                else:
                    # Create a dialog
                    dialog = ft.AlertDialog(
                        title=ft.Text("Внимание"),
                        content=ft.Text("Пожалуйста, выберите один параметр: слова или цифры."),
                        actions_alignment=ft.MainAxisAlignment.END,
                    )
                    
                    # Show the dialog
                    page.dialog = dialog
                    dialog.open = True
                    page.update()
                    
                
        return ' '.join(full_text)
    page.update()

    def on_button_ok_click(e):
      
        # Function for exporting words to .xlsx file
        def export_words_to_xlsx(text, xlsx_file):
            words = text.split(user_split.value)
            words = [word.strip() for word in words if word.strip()]  # Remove empty strings and whitespace
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_file)
            ws = wb.active
            
            if row_checkbox.value == 1:  # Fill in row
                if row_number.value:
                    row = int(row_number.value)  # Use the specified row number
                else:
                    row = ws.max_row + 1  # Find the first empty row
                for i, word in enumerate(words):
                    ws.cell(row=row, column=i+1, value=word)
            elif row_checkbox.value == 0:  # Fill in column
                if column_number.value:
                    col = int(column_number.value)  # Use the specified column number
                else:
                    col = ws.max_column + 1  # Find the first empty column
                for i, word in enumerate(words):
                    ws.cell(row=i+1, column=col, value=word)
            else:
                # Handle unexpected value of row_checkbox.value
                raise ValueError(f"Unexpected value of row_checkbox: {row_checkbox.value}")
            
            if text:  # Check if text was successfully extracted
                wb.save(xlsx_file)

                # Create and show a Snackbar to alert that the file was exported
                page.snack_bar = ft.SnackBar(content=ft.Text("Данные успешно добавлены в файл!"))
                page.snack_bar.open = True
                page.update()
            else:
                # Create and show a Snackbar to alert that there was an error
                page.snack_bar = ft.SnackBar(content=ft.Text("Ошибка! Не удалось извлечь текст из документа."))
                page.snack_bar.open = True
                page.update()

        # Path to the source .docx file
        docx_file = selected_docx_path

        # Path to the output .xlsx file
        xlsx_file = selected_xlsx_path

        # Extract text from .docx file
        text = extract_text_from_docx(docx_file)

        # Export words to .xlsx file
        export_words_to_xlsx(text, xlsx_file)

    page.add(ft.Text('Параметры разделения, стандартно ", " запятая и пробел'))
    user_split = ft.TextField(value=", ", width=150, height=35, border_color=ft.colors.BLUE)
    page.add(ft.Row(
        [
            user_split
        ]
    )
    )
    
    word_checkbox = ft.Checkbox("Оставить только слова, без цифр", value=False)
    number_checkbox = ft.Checkbox("Оставить только цифры, без слов", value=False)
    page.add(ft.Row(
        [
            word_checkbox,
            number_checkbox
        ]
    )
    )


    def update_rowcol_number_field(e):
        row_number.disabled = not row_checkbox.value
        column_number.disabled = row_checkbox.value
        row_number.value = ""
        column_number.value = ""
        page.update()

    row_checkbox = ft.Checkbox("Заполнить ячейки в строку", value=False, on_change=update_rowcol_number_field)
    page.add(ft.Row(
        [
            row_checkbox
        ]
    )
    )

    row_number = ft.TextField(value="", width=100, height=35, border_color=ft.colors.BLUE, disabled=not row_checkbox.value)
    page.add(ft.Row(
        [
            row_number,
            ft.Text("Номер строки (пустое значение = после последней занятой строки)")
        ]
    )
    )
    column_number = ft.TextField(value="", width=100, height=35, border_color=ft.colors.BLUE, disabled=row_checkbox.value)
    page.add(ft.Row(
        [
            column_number,
            ft.Text("Номер столбца (пустое значение = после последнего занятого солбца)")
        ]
    )
    )

    page.add(ft.Row(
        [
            ft.IconButton(ft.icons.ADS_CLICK, icon_size=35,on_click=on_button_ok_click),
            ft.Text("Нажмите для записи", color=ft.colors.RED)
        ]
    )
    )

ft.app(target=main)